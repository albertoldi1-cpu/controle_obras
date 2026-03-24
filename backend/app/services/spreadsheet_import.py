"""Importação em lote via planilha Excel (.xlsx ou .xls) para lançamentos físicos e financeiros."""
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any, List, Tuple

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import DailyEntry, FinancialDailyPlan, FinancialDailyProduction, FinancialTeam, Project, Stage


def _parse_date(s: str) -> date:
    s = (s or "").strip()
    if not s:
        raise ValueError("data vazia")
    parts = s.replace("/", "-").split("-")
    if len(parts) == 3 and len(parts[0]) == 4:
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
    elif len(parts) == 3:
        d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
        if y < 100:
            y += 2000
    else:
        raise ValueError(f"data inválida: {s}")
    return date(y, m, d)


def _float_cell(s: str, default: float = 0.0) -> float:
    t = (s or "").strip().replace(",", ".")
    if not t:
        return default
    return float(t)


def _int_cell(s: str) -> int:
    t = (s or "").strip().replace(",", ".")
    if not t:
        raise ValueError("número vazio")
    return int(float(t))


def _cell_to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(v)
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _read_rows_xlsx(data: bytes) -> List[List[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [_cell_to_str(c) for c in row]
        while cells and cells[-1] == "":
            cells.pop()
        if cells:
            rows.append(cells)
    return rows


def _read_rows_xls(data: bytes) -> List[List[str]]:
    import xlrd
    from xlrd.xldate import xldate_as_datetime

    book = xlrd.open_workbook(file_contents=data)
    sh = book.sheet_by_index(0)
    rows: List[List[str]] = []
    for r in range(sh.nrows):
        line: List[str] = []
        for c in range(sh.ncols):
            cell = sh.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xldate_as_datetime(cell.value, book.datemode)
                    line.append(dt.date().isoformat())
                except Exception:
                    line.append(str(cell.value))
            else:
                v = cell.value
                line.append("" if v == "" else _cell_to_str(v))
        while line and line[-1] == "":
            line.pop()
        if line:
            rows.append(line)
    return rows


def parse_upload_to_rows(content: bytes, filename: str) -> List[List[str]]:
    fn = (filename or "").lower()
    if fn.endswith(".xlsx"):
        return _read_rows_xlsx(content)
    if fn.endswith(".xls") and not fn.endswith(".xlsx"):
        try:
            return _read_rows_xls(content)
        except ImportError as e:
            raise ValueError(
                "Arquivo .xls requer o pacote xlrd no servidor. Salve como .xlsx ou reinstale dependências (pip install -r requirements.txt)."
            ) from e
    raise ValueError("Envie um arquivo .xlsx ou .xls (Excel).")


def _row_nonempty(row: List[str]) -> bool:
    return bool(row) and any((c or "").strip() for c in row)


def import_entries_planned(db: Session, project_id: int, rows: List[List[str]]) -> Tuple[int, List[str]]:
    """Linha 1 = cabeçalho. Colunas: stage_id | day | planned_optimistic | planned_pessimistic."""
    errors: List[str] = []
    stage_ids = set(db.scalars(select(Stage.id).where(Stage.project_id == project_id)).all())
    if not stage_ids:
        return 0, ["Projeto sem etapas cadastradas."]
    if not rows:
        return 0, ["Planilha vazia."]
    n = 0
    for i, row in enumerate(rows[1:], start=2):
        if not _row_nonempty(row):
            continue
        try:
            if len(row) < 4:
                raise ValueError("menos de 4 colunas preenchidas")
            sid = _int_cell(row[0])
            if sid not in stage_ids:
                raise ValueError(f"etapa {sid} não pertence ao projeto")
            d = _parse_date(row[1])
            opt = _float_cell(row[2])
            pes = _float_cell(row[3])
            existing = db.scalar(select(DailyEntry).where(DailyEntry.stage_id == sid, DailyEntry.day == d))
            if existing:
                existing.planned_optimistic = opt
                existing.planned_pessimistic = pes
            else:
                db.add(
                    DailyEntry(
                        stage_id=sid,
                        day=d,
                        planned_optimistic=opt,
                        planned_pessimistic=pes,
                        executed=0.0,
                        execution_note=None,
                    )
                )
            n += 1
        except Exception as e:
            errors.append(f"Linha {i}: {e}")
    if n:
        db.commit()
    return n, errors


def import_entries_executed(db: Session, project_id: int, rows: List[List[str]]) -> Tuple[int, List[str]]:
    """Linha 1 = cabeçalho. Colunas: stage_id | day | executed | execution_note (opcional)."""
    errors: List[str] = []
    stage_ids = set(db.scalars(select(Stage.id).where(Stage.project_id == project_id)).all())
    if not stage_ids:
        return 0, ["Projeto sem etapas cadastradas."]
    if not rows:
        return 0, ["Planilha vazia."]
    n = 0
    for i, row in enumerate(rows[1:], start=2):
        if not _row_nonempty(row):
            continue
        try:
            if len(row) < 3:
                raise ValueError("menos de 3 colunas preenchidas")
            sid = _int_cell(row[0])
            if sid not in stage_ids:
                raise ValueError(f"etapa {sid} não pertence ao projeto")
            d = _parse_date(row[1])
            ex = _float_cell(row[2])
            note = (row[3].strip() if len(row) > 3 else "") or None
            existing = db.scalar(select(DailyEntry).where(DailyEntry.stage_id == sid, DailyEntry.day == d))
            if existing:
                existing.executed = ex
                if note is not None:
                    existing.execution_note = note
            else:
                db.add(
                    DailyEntry(
                        stage_id=sid,
                        day=d,
                        planned_optimistic=0.0,
                        planned_pessimistic=0.0,
                        executed=ex,
                        execution_note=note,
                    )
                )
            n += 1
        except Exception as e:
            errors.append(f"Linha {i}: {e}")
    if n:
        db.commit()
    return n, errors


def _team_in_project(db: Session, project_id: int, team_id: int) -> None:
    t = db.get(FinancialTeam, team_id)
    if not t or t.project_id != project_id:
        raise ValueError(f"equipe {team_id} inválida")


def import_financial_plans(db: Session, project_id: int, rows: List[List[str]]) -> Tuple[int, List[str]]:
    """Linha 1 = cabeçalho. Colunas: day | team_id | daily_target_brl | daily_planning_brl (opcional)."""
    if not db.get(Project, project_id):
        return 0, ["Projeto não encontrado."]
    errors: List[str] = []
    if not rows:
        return 0, ["Planilha vazia."]
    n = 0
    for i, row in enumerate(rows[1:], start=2):
        if not _row_nonempty(row):
            continue
        try:
            if len(row) < 3:
                raise ValueError("menos de 3 colunas preenchidas")
            d = _parse_date(row[0])
            tid = _int_cell(row[1])
            _team_in_project(db, project_id, tid)
            tgt = _float_cell(row[2])
            pln = _float_cell(row[3], 0.0) if len(row) > 3 else 0.0
            existing = db.scalar(
                select(FinancialDailyPlan).where(
                    FinancialDailyPlan.project_id == project_id,
                    FinancialDailyPlan.day == d,
                    FinancialDailyPlan.team_id == tid,
                )
            )
            if existing:
                existing.daily_target_brl = tgt
                existing.daily_planning_brl = pln
            else:
                db.add(
                    FinancialDailyPlan(
                        project_id=project_id,
                        day=d,
                        team_id=tid,
                        daily_target_brl=tgt,
                        daily_planning_brl=pln,
                    )
                )
            n += 1
        except Exception as e:
            errors.append(f"Linha {i}: {e}")
    if n:
        db.commit()
    return n, errors


def import_financial_production(db: Session, project_id: int, rows: List[List[str]]) -> Tuple[int, List[str]]:
    """Linha 1 = cabeçalho. Colunas: day | team_id | produced_value_brl | observation (opcional)."""
    if not db.get(Project, project_id):
        return 0, ["Projeto não encontrado."]
    errors: List[str] = []
    if not rows:
        return 0, ["Planilha vazia."]
    n = 0
    for i, row in enumerate(rows[1:], start=2):
        if not _row_nonempty(row):
            continue
        try:
            if len(row) < 3:
                raise ValueError("menos de 3 colunas preenchidas")
            d = _parse_date(row[0])
            tid = _int_cell(row[1])
            _team_in_project(db, project_id, tid)
            val = _float_cell(row[2])
            obs = (row[3].strip() if len(row) > 3 else "") or None
            existing = db.scalar(
                select(FinancialDailyProduction).where(
                    FinancialDailyProduction.project_id == project_id,
                    FinancialDailyProduction.day == d,
                    FinancialDailyProduction.team_id == tid,
                )
            )
            if existing:
                existing.produced_value_brl = val
                existing.observation = obs
            else:
                db.add(
                    FinancialDailyProduction(
                        project_id=project_id,
                        day=d,
                        team_id=tid,
                        produced_value_brl=val,
                        observation=obs,
                    )
                )
            n += 1
        except Exception as e:
            errors.append(f"Linha {i}: {e}")
    if n:
        db.commit()
    return n, errors


def import_from_upload(
    db: Session,
    project_id: int,
    content: bytes,
    filename: str,
    *,
    kind: str,
    scope: str,
) -> Tuple[int, List[str]]:
    """
    scope: 'entries' | 'financial'
    kind: entries -> 'planned' | 'executed'; financial -> 'plans' | 'production'
    """
    rows = parse_upload_to_rows(content, filename)
    if scope == "entries":
        if kind == "planned":
            return import_entries_planned(db, project_id, rows)
        if kind == "executed":
            return import_entries_executed(db, project_id, rows)
        raise ValueError("kind inválido para lançamentos")
    if scope == "financial":
        if kind == "plans":
            return import_financial_plans(db, project_id, rows)
        if kind == "production":
            return import_financial_production(db, project_id, rows)
        raise ValueError("kind inválido para financeiro")
    raise ValueError("scope inválido")
