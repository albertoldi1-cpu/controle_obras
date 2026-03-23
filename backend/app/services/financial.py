"""Painel financeiro: meta total das equipes (planejado) × produzido, curva acumulada e farol diário."""
from __future__ import annotations

from collections import defaultdict
from datetime import date
from typing import List, Optional, Tuple

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models import FinancialDailyPlan, FinancialDailyProduction, FinancialTeam, Project
from app.schemas import (
    Farol,
    FinancialFarolDayRow,
    FinancialPhysicalComparisonOut,
    FinancialPhysicalComparisonPoint,
    FinancialPhysicalComparisonSummary,
    FinancialPanelDashboardOut,
    FinancialPanelFiltersOut,
    FinancialPanelSeriesPoint,
    FinancialPanelSummary,
    FinancialTeamBriefOut,
)

# Banda “pessimista” para o farol (mesma ideia do avanço físico)
FIN_PEAK = 0.85


def _farol(planned: float, produced: float) -> Farol:
    if planned <= 0 and produced <= 0:
        return "green"
    if planned <= 0:
        return "green"
    if produced >= planned:
        return "green"
    if produced >= planned * FIN_PEAK:
        return "yellow"
    return "red"


def _filter_team_id(row_team_id: int, filt: Optional[int]) -> bool:
    if filt is None:
        return True
    return int(row_team_id) == int(filt)


def _physical_executed_pct_at_day(
    stages: list,
    day: date,
    date_from: Optional[date],
    date_to: Optional[date],
) -> float:
    """% avanço físico acumulado até `day` (mesma lógica do painel físico), respeitando filtros de data."""
    acc = 0.0
    for st in stages:
        q = float(st.total_quantity or 0.0)
        if q <= 0:
            continue
        executed = 0.0
        for e in st.entries:
            if date_from and e.day < date_from:
                continue
            if date_to and e.day > date_to:
                continue
            if e.day <= day:
                executed += float(e.executed or 0.0)
        acc += float(st.weight) * min(executed / q, 1.0)
    return round(acc * 100.0, 3)


def build_financial_panel_dashboard(
    db: Session,
    project: Project,
    date_from: Optional[date],
    date_to: Optional[date],
    team_filter: Optional[int],
) -> FinancialPanelDashboardOut:
    pid = project.id
    q_plans = select(FinancialDailyPlan).where(FinancialDailyPlan.project_id == pid)
    q_prod = select(FinancialDailyProduction).where(FinancialDailyProduction.project_id == pid)
    if date_from:
        q_plans = q_plans.where(FinancialDailyPlan.day >= date_from)
        q_prod = q_prod.where(FinancialDailyProduction.day >= date_from)
    if date_to:
        q_plans = q_plans.where(FinancialDailyPlan.day <= date_to)
        q_prod = q_prod.where(FinancialDailyProduction.day <= date_to)

    plans = list(db.scalars(q_plans).all())
    prods = list(db.scalars(q_prod).all())

    team_rows = list(
        db.scalars(select(FinancialTeam).where(FinancialTeam.project_id == pid).order_by(FinancialTeam.name)).all()
    )
    teams_brief = [FinancialTeamBriefOut.model_validate(t) for t in team_rows]

    plans_f = [p for p in plans if _filter_team_id(p.team_id, team_filter)]
    prods_f = [p for p in prods if _filter_team_id(p.team_id, team_filter)]

    by_day_team_meta: dict[date, dict[int, float]] = defaultdict(dict)
    by_day_produced: dict[date, float] = defaultdict(float)
    by_day_team_ids: dict[date, set[int]] = defaultdict(set)

    for p in plans_f:
        by_day_team_meta[p.day][p.team_id] = float(p.daily_target_brl)
    for p in prods_f:
        by_day_produced[p.day] += float(p.produced_value_brl)
        if float(p.produced_value_brl) > 0:
            by_day_team_ids[p.day].add(p.team_id)

    by_day_meta_total: dict[date, float] = {}
    all_days = sorted(set(by_day_team_meta.keys()) | set(by_day_produced.keys()))
    for d in all_days:
        # Meta Total das Equipes:
        # soma das metas das equipes que efetivamente produziram no dia.
        produced_teams = by_day_team_ids.get(d, set())
        day_meta = 0.0
        if produced_teams:
            day_meta = sum(by_day_team_meta.get(d, {}).get(tid, 0.0) for tid in produced_teams)
        by_day_meta_total[d] = day_meta

    cum_p = 0.0
    cum_e = 0.0
    stages_list = list(project.stages)
    obra_total = float(project.obra_total_value_brl or 0.0)
    series: List[FinancialPanelSeriesPoint] = []
    farol_rows: List[FinancialFarolDayRow] = []

    for d in all_days:
        dp = by_day_meta_total.get(d, 0.0)
        dr = by_day_produced.get(d, 0.0)
        cum_p += dp
        cum_e += dr
        phys_pct = _physical_executed_pct_at_day(stages_list, d, date_from, date_to)
        prod_pct = (
            round(min((cum_e / obra_total) * 100.0, 100.0), 3) if obra_total > 1e-9 else 0.0
        )
        series.append(
            FinancialPanelSeriesPoint(
                day=d,
                daily_planned_brl=dp,
                daily_produced_brl=dr,
                cumulative_planned_brl=cum_p,
                cumulative_produced_brl=cum_e,
                physical_executed_pct=phys_pct,
                productive_advance_pct=prod_pct,
            )
        )
        farol_rows.append(
            FinancialFarolDayRow(
                day=d,
                planned_brl=dp,
                produced_brl=dr,
                teams_count=len(by_day_team_ids.get(d, set())),
                farol=_farol(dp, dr),
            )
        )

    total_p = sum(by_day_meta_total.values())
    total_r = sum(by_day_produced.values())
    dev: Optional[float] = None
    if total_p > 0:
        dev = (total_r - total_p) / total_p * 100.0
    last_day = max(all_days) if all_days else None

    summary = FinancialPanelSummary(
        total_planned_brl=total_p,
        total_produced_brl=total_r,
        deviation_pct=dev,
        last_data_day=last_day,
    )
    filters = FinancialPanelFiltersOut(
        date_from=date_from,
        date_to=date_to,
        team_id=team_filter,
    )

    return FinancialPanelDashboardOut(
        project_id=project.id,
        project_name=project.name,
        obra_total_value_brl=float(project.obra_total_value_brl)
        if project.obra_total_value_brl is not None
        else None,
        filters=filters,
        summary=summary,
        series=series,
        farol_days=farol_rows,
        teams=teams_brief,
    )


def financial_excel_bytes(
    db: Session,
    project: Project,
    date_from: Optional[date],
    date_to: Optional[date],
    team_filter: Optional[int],
) -> Tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    dash = build_financial_panel_dashboard(db, project, date_from, date_to, team_filter)
    wb = Workbook()
    # Painel
    ws0 = wb.active
    ws0.title = "Painel"
    ws0.append(["Projeto", dash.project_name])
    ws0.append(["Filtro data de", str(dash.filters.date_from or "")])
    ws0.append(["Filtro data até", str(dash.filters.date_to or "")])
    ws0.append(["Filtro equipe (id)", dash.filters.team_id or "Todas"])
    ws0.append([])
    ws0.append(["Meta Total das Equipes (R$)", dash.summary.total_planned_brl])
    ws0.append(["Total produzido (R$)", dash.summary.total_produced_brl])
    ws0.append(["Desvio %", dash.summary.deviation_pct if dash.summary.deviation_pct is not None else ""])
    ws0.append(
        [
            "Valor total da obra (R$)",
            dash.obra_total_value_brl if dash.obra_total_value_brl is not None else "",
        ]
    )
    ws0.append([])
    ws0.append(
        [
            "Data",
            "Meta Total das Equipes (dia)",
            "Produzido (dia)",
            "Acum. Meta Total das Equipes",
            "Acum. produzido",
            "Avanço físico (%)",
            "Avanço produtivo (%)",
        ]
    )
    for p in dash.series:
        ws0.append(
            [
                p.day.isoformat(),
                p.daily_planned_brl,
                p.daily_produced_brl,
                p.cumulative_planned_brl,
                p.cumulative_produced_brl,
                p.physical_executed_pct,
                p.productive_advance_pct,
            ]
        )
    ws0.append([])
    ws0.append(["Farol por dia"])
    ws0.append(["Data", "Meta Total das Equipes", "Produzido", "Qtd equipes ativas", "Farol"])
    for f in dash.farol_days:
        ws0.append([f.day.isoformat(), f.planned_brl, f.produced_brl, f.teams_count, f.farol])

    # Planejamento
    ws1 = wb.create_sheet("Metas Equipes")
    ws1.append(["Data", "Equipe", "Tipo", "Meta da equipe (R$)"])
    q = (
        select(FinancialDailyPlan)
        .where(FinancialDailyPlan.project_id == project.id)
        .order_by(FinancialDailyPlan.day, FinancialDailyPlan.team_id)
    )
    if date_from:
        q = q.where(FinancialDailyPlan.day >= date_from)
    if date_to:
        q = q.where(FinancialDailyPlan.day <= date_to)
    for row in db.scalars(q).all():
        if not _filter_team_id(row.team_id, team_filter):
            continue
        team = db.get(FinancialTeam, row.team_id)
        nm = team.name if team else "—"
        tt = team.team_type if team else "—"
        ws1.append([row.day.isoformat(), nm, tt, row.daily_target_brl])

    # Lançamentos produtividade
    ws2 = wb.create_sheet("Lançamentos")
    ws2.append(["Data", "Equipe", "Tipo", "Meta da equipe (R$)", "Valor produzido (R$)", "Observações"])
    plans_map: dict[tuple[date, int], float] = {}
    q_plan_map = select(FinancialDailyPlan).where(FinancialDailyPlan.project_id == project.id)
    if date_from:
        q_plan_map = q_plan_map.where(FinancialDailyPlan.day >= date_from)
    if date_to:
        q_plan_map = q_plan_map.where(FinancialDailyPlan.day <= date_to)
    for p in db.scalars(q_plan_map).all():
        plans_map[(p.day, p.team_id)] = float(p.daily_target_brl)
    q2 = (
        select(FinancialDailyProduction)
        .where(FinancialDailyProduction.project_id == project.id)
        .order_by(FinancialDailyProduction.day)
    )
    if date_from:
        q2 = q2.where(FinancialDailyProduction.day >= date_from)
    if date_to:
        q2 = q2.where(FinancialDailyProduction.day <= date_to)
    for row in db.scalars(q2).all():
        if not _filter_team_id(row.team_id, team_filter):
            continue
        team = db.get(FinancialTeam, row.team_id)
        nm = team.name if team else "—"
        tt = team.team_type if team else "—"
        ws2.append(
            [
                row.day.isoformat(),
                nm,
                tt,
                plans_map.get((row.day, row.team_id), 0.0),
                row.produced_value_brl,
                row.observation or "",
            ]
        )

    for ws in wb.worksheets:
        for c in ws[1]:
            if c.value:
                c.font = Font(bold=True)

    import io

    bio = io.BytesIO()
    wb.save(bio)
    fname = f"painel-financeiro-{project.id}.xlsx"
    return bio.getvalue(), fname


def build_financial_physical_comparison(
    db: Session,
    project: Project,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> FinancialPhysicalComparisonOut:
    stages = list(project.stages)
    phys_by_day_pct: dict[date, float] = {}
    phys_opt_by_day_pct: dict[date, float] = {}
    phys_pes_by_day_pct: dict[date, float] = {}
    all_phys_days: set[date] = set()
    for st in stages:
        for e in st.entries:
            if (date_from and e.day < date_from) or (date_to and e.day > date_to):
                continue
            all_phys_days.add(e.day)

    sorted_phys_days = sorted(all_phys_days)
    for d in sorted_phys_days:
        acc = 0.0
        acc_opt = 0.0
        acc_pes = 0.0
        for st in stages:
            q = float(st.total_quantity or 0.0)
            if q <= 0:
                continue
            executed = 0.0
            planned_opt = 0.0
            planned_pes = 0.0
            for e in st.entries:
                if (date_from and e.day < date_from) or (date_to and e.day > date_to):
                    continue
                if e.day <= d:
                    executed += float(e.executed or 0.0)
                    planned_opt += float(e.planned_optimistic or 0.0)
                    planned_pes += float(e.planned_pessimistic or 0.0)
            acc += float(st.weight) * min(executed / q, 1.0)
            acc_opt += float(st.weight) * min(planned_opt / q, 1.0)
            acc_pes += float(st.weight) * min(planned_pes / q, 1.0)
        phys_by_day_pct[d] = round(acc * 100, 3)
        phys_opt_by_day_pct[d] = round(acc_opt * 100, 3)
        phys_pes_by_day_pct[d] = round(acc_pes * 100, 3)

    q_prod_rows = select(FinancialDailyProduction).where(FinancialDailyProduction.project_id == project.id)
    if date_from:
        q_prod_rows = q_prod_rows.where(FinancialDailyProduction.day >= date_from)
    if date_to:
        q_prod_rows = q_prod_rows.where(FinancialDailyProduction.day <= date_to)
    prod_rows = list(db.scalars(q_prod_rows).all())

    produced_by_day: dict[date, float] = defaultdict(float)
    for r in prod_rows:
        produced_by_day[r.day] += float(r.produced_value_brl or 0.0)

    q_plan_days = select(func.count(func.distinct(FinancialDailyPlan.day))).where(
        FinancialDailyPlan.project_id == project.id
    )
    if date_from:
        q_plan_days = q_plan_days.where(FinancialDailyPlan.day >= date_from)
    if date_to:
        q_plan_days = q_plan_days.where(FinancialDailyPlan.day <= date_to)
    n_plan_days = int(db.scalar(q_plan_days) or 0)

    obra_total = float(project.obra_total_value_brl or 0.0)
    daily_obra_ref = (obra_total / n_plan_days) if n_plan_days > 0 and obra_total > 1e-9 else 0.0

    all_days = sorted(set(sorted_phys_days) | set(produced_by_day.keys()))
    points: List[FinancialPhysicalComparisonPoint] = []
    cum_prod = 0.0
    last_phys = 0.0
    last_phys_opt = 0.0
    last_phys_pes = 0.0
    ratio_brl_per_physical_point = 0.0
    if all_days:
        final_phys = phys_by_day_pct.get(all_days[-1], 0.0)
        if final_phys > 1e-9:
            ratio_brl_per_physical_point = sum(produced_by_day.values()) / final_phys
    for d in all_days:
        if d in phys_by_day_pct:
            last_phys = phys_by_day_pct[d]
        if d in phys_opt_by_day_pct:
            last_phys_opt = phys_opt_by_day_pct[d]
        if d in phys_pes_by_day_pct:
            last_phys_pes = phys_pes_by_day_pct[d]
        day_prod = produced_by_day.get(d, 0.0)
        cum_prod += day_prod
        forecast_opt_cum = last_phys_opt * ratio_brl_per_physical_point
        forecast_pes_cum = last_phys_pes * ratio_brl_per_physical_point
        prev_opt = points[-1].optimistic_productive_forecast_brl if points else 0.0
        prev_pes = points[-1].pessimistic_productive_forecast_brl if points else 0.0
        points.append(
            FinancialPhysicalComparisonPoint(
                day=d,
                physical_executed_pct=last_phys,
                produced_value_brl=day_prod,
                optimistic_productive_forecast_brl=max(forecast_opt_cum - prev_opt, 0.0),
                pessimistic_productive_forecast_brl=max(forecast_pes_cum - prev_pes, 0.0),
                cumulative_produced_value_brl=cum_prod,
                daily_obra_reference_brl=daily_obra_ref,
            )
        )

    summary = FinancialPhysicalComparisonSummary(
        last_day=all_days[-1] if all_days else None,
        physical_executed_pct=last_phys,
        total_produced_brl=cum_prod,
        obra_total_value_brl=obra_total,
        planned_financial_days_count=n_plan_days,
        daily_obra_reference_brl=daily_obra_ref,
    )
    return FinancialPhysicalComparisonOut(
        project_id=project.id,
        project_name=project.name,
        points=points,
        summary=summary,
    )


def financial_physical_comparison_excel_bytes(
    db: Session,
    project: Project,
    date_from: Optional[date],
    date_to: Optional[date],
) -> Tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    import io

    comp = build_financial_physical_comparison(db, project, date_from, date_to)
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativo"
    ws.append(["Projeto", comp.project_name])
    ws.append(["Filtro data de", str(date_from or "")])
    ws.append(["Filtro data até", str(date_to or "")])
    ws.append([])
    ws.append(["Avanço físico executado (%)", comp.summary.physical_executed_pct])
    ws.append(["Total produzido (R$)", comp.summary.total_produced_brl])
    ws.append(["Valor total da obra (R$)", comp.summary.obra_total_value_brl])
    ws.append(["Dias com planejamento financeiro (distintos)", comp.summary.planned_financial_days_count])
    ws.append(["Referência diária obra (R$)", comp.summary.daily_obra_reference_brl])
    ws.append([])
    ws.append(
        [
            "Data",
            "Executado físico (%)",
            "Produzido (R$ dia)",
            "Previsão produtiva otimista (R$ dia)",
            "Previsão produtiva pessimista (R$ dia)",
            "Produzido acumulado (R$)",
            "Referência diária obra (R$)",
        ]
    )
    for p in comp.points:
        ws.append(
            [
                p.day.isoformat(),
                p.physical_executed_pct,
                p.produced_value_brl,
                p.optimistic_productive_forecast_brl,
                p.pessimistic_productive_forecast_brl,
                p.cumulative_produced_value_brl,
                p.daily_obra_reference_brl,
            ]
        )
    for c in ws[1]:
        if c.value:
            c.font = Font(bold=True)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue(), f"comparativo-fisico-produtivo-{project.id}.xlsx"
