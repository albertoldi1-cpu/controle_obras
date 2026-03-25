"""Avanço financeiro da obra: planejado (import ou somas do app) × avanço produtivo (%)."""
from __future__ import annotations

import re
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import FinancialDailyPlan, FinancialDailyProduction, FinancialObraPlanDaily, Project
from app.schemas import ObraFinancialAdvanceOut, ObraFinancialAdvancePoint


def _cell_to_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            return from_excel(v).date()
        except Exception:
            return None
    return None


def _float_cell(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", ".")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _norm_sheet_title(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _is_likely_cumulative_row(vals: List[float]) -> bool:
    if len(vals) < 4:
        return False
    return all(vals[i] <= vals[i + 1] + 1e-6 for i in range(len(vals) - 1)) and vals[-1] > vals[0] + 1e-6


def parse_avanco_financeiro_xlsx(content: bytes) -> Dict[date, float]:
    """
    Lê a folha «AVANÇO FINANCEIRO» (ou nome parecido): bloco cenário otimista com datas na linha de cabeçalho.
    Soma linhas de detalhe por coluna de data; ignora linhas monótonas crescentes (curva acumulada).
    """
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    try:
        ws = None
        for name in wb.sheetnames:
            n = _norm_sheet_title(name)
            if "avan" in n and "financeir" in n:
                ws = wb[name]
                break
        if ws is None:
            for name in wb.sheetnames:
                if "financeir" in _norm_sheet_title(name):
                    ws = wb[name]
                    break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    header_idx: Optional[int] = None
    date_cols: List[Tuple[int, date]] = []

    def _collect_date_cols(row) -> List[Tuple[int, date]]:
        out: List[Tuple[int, date]] = []
        for j in range(1, len(row)):
            d = _cell_to_date(row[j])
            if d:
                out.append((j, d))
        return out

    for ri, row in enumerate(rows):
        if not row:
            continue
        c0 = str(row[0] or "").upper()
        if "PESSIMISTA" in c0 and header_idx is not None:
            break
        cols = _collect_date_cols(row)
        if "OTIMISTA" in c0 and len(row) > 1 and _cell_to_date(row[1]) and len(cols) >= 3:
            date_cols = cols
            header_idx = ri
            break

    if header_idx is None:
        for ri, row in enumerate(rows):
            if not row:
                continue
            cols = _collect_date_cols(row)
            if len(cols) >= 8:
                date_cols = cols
                header_idx = ri
                break

    if header_idx is None or not date_cols:
        raise ValueError(
            "Não foi possível localizar o bloco com datas na linha de cabeçalho (folha «AVANÇO FINANCEIRO»)."
        )

    daily: Dict[date, float] = defaultdict(float)

    for ri in range(header_idx + 1, len(rows)):
        row = rows[ri]
        if not row:
            continue
        c0 = str(row[0] or "").upper()
        if "PESSIMISTA" in c0:
            break

        vals: List[float] = []
        for j, _dt in date_cols:
            if j < len(row):
                fv = _float_cell(row[j])
                if fv is not None:
                    vals.append(fv)
        if len(vals) < 3:
            continue
        if _is_likely_cumulative_row(vals):
            continue

        for j, dt in date_cols:
            if j < len(row):
                fv = _float_cell(row[j])
                if fv is not None and abs(fv) > 1e-12:
                    daily[dt] += fv

    if not daily:
        raise ValueError("Nenhum valor diário encontrado no bloco otimista (verifique o formato da planilha).")

    return dict(daily)


def replace_project_obra_plan(db: Session, project_id: int, daily: Dict[date, float]) -> int:
    db.query(FinancialObraPlanDaily).filter(FinancialObraPlanDaily.project_id == project_id).delete(
        synchronize_session=False
    )
    n = 0
    for d, inc in sorted(daily.items()):
        db.add(
            FinancialObraPlanDaily(
                project_id=project_id,
                day=d,
                planned_increment_brl=float(inc),
            )
        )
        n += 1
    db.commit()
    return n


def _planned_daily_from_app(db: Session, project_id: int) -> Dict[date, float]:
    plans = list(db.scalars(select(FinancialDailyPlan).where(FinancialDailyPlan.project_id == project_id)).all())
    by_day: Dict[date, float] = defaultdict(float)

    def _planning_brl(row: FinancialDailyPlan) -> float:
        v = getattr(row, "daily_planning_brl", None)
        return float(v) if v is not None else 0.0

    for p in plans:
        pln = _planning_brl(p)
        tgt = float(p.daily_target_brl or 0.0)
        by_day[p.day] += pln if pln > 1e-9 else tgt
    return dict(by_day)


def _productive_pct_by_day(db: Session, project_id: int, obra_total: float) -> Dict[date, float]:
    plans_list = list(db.scalars(select(FinancialDailyPlan).where(FinancialDailyPlan.project_id == project_id)).all())
    prods_list = list(
        db.scalars(select(FinancialDailyProduction).where(FinancialDailyProduction.project_id == project_id)).all()
    )
    by_day_produced: Dict[date, float] = defaultdict(float)
    for p in prods_list:
        by_day_produced[p.day] += float(p.produced_value_brl or 0.0)

    by_day_team_meta: Dict[date, Dict[int, float]] = defaultdict(dict)
    by_day_team_planning: Dict[date, Dict[int, float]] = defaultdict(dict)
    by_day_team_ids: Dict[date, set] = defaultdict(set)

    def _planning_brl(row: FinancialDailyPlan) -> float:
        v = getattr(row, "daily_planning_brl", None)
        return float(v) if v is not None else 0.0

    for p in plans_list:
        by_day_team_meta[p.day][p.team_id] = float(p.daily_target_brl or 0.0)
        by_day_team_planning[p.day][p.team_id] = _planning_brl(p)
    for p in prods_list:
        if float(p.produced_value_brl or 0.0) > 0:
            by_day_team_ids[p.day].add(p.team_id)

    fin_days = sorted(set(by_day_team_meta.keys()) | set(by_day_produced.keys()))
    cum_e = 0.0
    out: Dict[date, float] = {}
    for d in fin_days:
        cum_e += by_day_produced.get(d, 0.0)
        out[d] = round(min((cum_e / obra_total) * 100.0, 100.0), 3) if obra_total > 1e-9 else 0.0
    return out


def build_obra_financial_advance(db: Session, project: Project) -> ObraFinancialAdvanceOut:
    pid = project.id
    obra_total = float(project.obra_total_value_brl or 0.0)

    productive_pct_by_day = _productive_pct_by_day(db, pid, obra_total)

    imported_rows = list(
        db.scalars(
            select(FinancialObraPlanDaily)
            .where(FinancialObraPlanDaily.project_id == pid)
            .order_by(FinancialObraPlanDaily.day)
        ).all()
    )
    imported_by_day = {r.day: float(r.planned_increment_brl or 0.0) for r in imported_rows}

    if imported_by_day:
        source: str = "imported"
        planned_daily = imported_by_day
    else:
        source = "app_plans"
        planned_daily = _planned_daily_from_app(db, pid)

    all_days_set = set(planned_daily.keys()) | set(productive_pct_by_day.keys())
    if not all_days_set:
        return ObraFinancialAdvanceOut(
            project_id=project.id,
            project_name=project.name,
            obra_total_value_brl=project.obra_total_value_brl,
            source_planned=source,  # type: ignore[arg-type]
            series=[],
        )

    all_days = sorted(all_days_set)
    series: List[ObraFinancialAdvancePoint] = []
    cum_plan_brl = 0.0
    last_prod_pct = 0.0
    for d in all_days:
        cum_plan_brl += planned_daily.get(d, 0.0)
        planned_pct = round(min((cum_plan_brl / obra_total) * 100.0, 100.0), 3) if obra_total > 1e-9 else 0.0
        if d in productive_pct_by_day:
            last_prod_pct = productive_pct_by_day[d]
        series.append(
            ObraFinancialAdvancePoint(
                day=d,
                planned_financial_pct=planned_pct,
                productive_advance_pct=last_prod_pct,
            )
        )

    return ObraFinancialAdvanceOut(
        project_id=project.id,
        project_name=project.name,
        obra_total_value_brl=project.obra_total_value_brl,
        source_planned=source,  # type: ignore[arg-type]
        series=series,
    )
